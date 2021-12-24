print('Import library....')
from os import listdir
from os.path import join
import os
import openpyxl as opx
from datetime import datetime
from openpyxl.styles import PatternFill
print('Import library DONE....')


 # mengakses directori file sumber
print('Get Directory....')
arr = os.getcwd()
ab = os.listdir('Sumber')
tanggal_jam = datetime.now().strftime("Tanggal %d-%m-%Y __Pukul %H-%M-%S")
print('Get Directory DONE....')


print('Workbook......')
new_wr = opx.Workbook()
sheet = new_wr.active
print('Workbook DONE......')


list_header = [
    "TANGGAL", "NAMA SEKOLAH", "NPSN", "ALAMAT SEKOLAH",
    "NAMA LENGKAP (TANPA GELAR)", "NIK / NO. PASSPORT (UNTUK WN)", "JENIS KELAMIN",
    "TEMPAT LAHIR", "TANGGAL LAHIR", "NAMA IBU KANDUNG", "ALAMAT JALAN",
    "RT", "RW", "NAMA DUSUN", "DESA / KELURAHAN", "KECAMATAN", "KODE POS", "AGAMA", "NPWP",
    "NAMA WAJIB PAJAK", "KEWARGANEGARAAN", "STATUS PERKAWINAN", "NAMA SUAMI / ISTRI",
    "NIP SUAMI / ISTRI", "PEKERJAAN SUAMI / ISTRI", "STATUS PERKAWINAN",
    "NIP", "NIY / NIGK", "NUPTK", "JENIS PTK", "SK PENGANGKATAN", "TMT PENGANGKAT",
    "LEMBAGA PENGANGKAT", "SK CPNS", "TMT PNS", "TMT PNS",
    "PANGKAT GOLONGAN", "SUMBER GAJI", "KARTU PEGAWAI",
    "KARTU ISTRI (KARIS) / KARTU SUAMI (KARSU)", "PUNYA LISENSI KEPALA SEKOLAH",
    "KEAHLIAN LABORATORIUM", "MAMPU MENANGANI KEBUTUHAN KHUSUS",
    "KEAHLIAN BRAILE", "KEAHLIAN BHS. ISYARAT",
    "NOMOR TELEPON RUMAH", "NOMOR HP", "EMAIL",
    "ID BANK", "NOMOR REKENING BANK", "REKENING ATAS NAMA",
    "NOMOR SURAT TUGAS", "TANGGAL SURAT TUGAS", "TMT TUGAS",
    "STATUS SEKOLAH UNTUK", "KELUAR KARENA", "TANGGAL KELUAR", 
]

sheet_riwayat_sertifikasi = 'Riwayat_Sertifikasi'
sheet_riwayat_pendidikan = 'Riwayat_Pendidikan'

print('Making Worksheet...')
for x in range(len(list_header)):
    sheet.cell(row=1, column=x+1).value = list_header[x]
     # sheet.cell(row=1, column=x+1).font = Font(bold=True, color='040CFF')
for jj in range(1, 3000):
    sheet.cell(row=4, column=jj).value = jj
abjad = ["A","B","C","D","E","F","G","H","I",
"J","K","L","M","N","O","P","Q","R","S","T","U","V",
"W","X","Y","Z",
"AA","BB","CC","DD","EE","FF",
"GG","HH","II","JJ","KK","LL","MM",
"NN","OO","PP","QQ","RR","SS","TT",
"UU","VV","WW","XX","YY","ZZ",

]

header_list_RS = []
header_list_RP = []
header_list_KO = []
header_list_AK = []
header_list_BE = []
header_list_BK = []
header_list_DK = []
header_list_KT = []
header_list_KS = []
header_list_TJ = []
header_list_TT = []
header_list_IN = []
header_list_PH = []
header_list_NT = []
header_list_RG = []
header_list_RK = []
header_list_RJPTK = []
header_list_KG = []
header_list_JF = []


sumber_data = os.path.join(arr, 'Data_MENTAH_JANGAN_DIHAPUS', 'F-PTK-JANGAN_DIHAPUS.xlsx')
sumber_data_worksheet = opx.load_workbook(sumber_data, data_only=True)

sheet_Riwayat_Sertifikasi=sumber_data_worksheet["Riwayat_Sertifikasi"]							
sheet_Riwayat_Pendidikan=sumber_data_worksheet["Riwayat_Pendidikan"]							
sheet_Kompetensi=sumber_data_worksheet["Kompetensi"]							
sheet_Anak=sumber_data_worksheet["Anak"]							
sheet_Beasiswa=sumber_data_worksheet["Beasiswa"]							
sheet_Buku_yang_pernah_ditulis=sumber_data_worksheet["Buku_yang_pernah_ditulis"]							
sheet_Diklat=sumber_data_worksheet["Diklat"]							
sheet_Karya_Tulis=sumber_data_worksheet["Karya_Tulis"]							
sheet_Kesejahteraan=sumber_data_worksheet["Kesejahteraan"]							
sheet_Tunjangan=sumber_data_worksheet["Tunjangan"]							
sheet_Tugas_Tambahan=sumber_data_worksheet["Tugas_Tambahan"]							
sheet_Inpassing_Non_PNS=sumber_data_worksheet["Inpassing_Non_PNS"]							
sheet_Penghargaan=sumber_data_worksheet["Penghargaan"]							
sheet_Nilai_Tes=sumber_data_worksheet["Nilai_Tes"]							
sheet_Riwayat_Gaji_Berkala=sumber_data_worksheet["Riwayat_Gaji_Berkala"]							
sheet_Riwayat_Karir_Guru=sumber_data_worksheet["Riwayat_Karir_Guru"]							
sheet_Riwayat_Jabatan_P_TK=sumber_data_worksheet["Riwayat_Jabatan_P_TK"]							
sheet_Riwayat_Kepangkatan_Golongan=sumber_data_worksheet["Riwayat_Kepangkatan_Golongan"]							
sheet_Riwayat_Jabatan_Fungsional=sumber_data_worksheet["Riwayat_Jabatan_Fungsional"]							



for data_RS in range(1, 9):
    for sheet_riwayat_sertifikasi in range(1, 7):
        dt_tb = sheet_Riwayat_Sertifikasi.cell(row=1, column=sheet_riwayat_sertifikasi).value
        header_list_RS.append(str(dt_tb) + "_" + str(abjad[data_RS-1]))
    for a in range(len(header_list_RS)):
        sheet.cell(row=1, column=a+58).value = header_list_RS[a]


for data_RP in range(1, 12):
    for sheet_riwayat_pendidikan in range(1, 11):
        dt_tb = sheet_Riwayat_Pendidikan.cell(row=1, column=sheet_riwayat_pendidikan).value
        header_list_RP.append(str(dt_tb) + "_" + str(abjad[data_RP-1]))
    for a in range(len(header_list_RP)):
        sheet.cell(row=1, column=a+100).value = header_list_RP[a]


for data_KO in range(1, 10):
    for sheet_kompetensi in range(1, 3):
        dt_tb = sheet_Kompetensi.cell(row=1, column=sheet_kompetensi).value
        header_list_KO.append(str(dt_tb) + "_" + str(abjad[data_KO-1]))
    for a in range(len(header_list_KO)):
        sheet.cell(row=1, column=a+210).value = header_list_KO[a]


for data_AK in range(1, 10):
    for sheet_anak in range(1, 9):
        dt_tb = sheet_Anak.cell(row=1, column=sheet_anak).value
        header_list_AK.append(str(dt_tb) + "_" + str(abjad[data_AK-1]))
    for a in range(len(header_list_AK)):
        sheet.cell(row=1, column=a+228).value = header_list_AK[a]


for data_BE in range(1, 10):
    for sheet_beasiswa in range(1, 6):
        dt_tb = sheet_Beasiswa.cell(row=1, column=sheet_beasiswa).value
        header_list_BE.append(str(dt_tb) + "_" + str(abjad[data_BE-1]))
    for a in range(len(header_list_BE)):
        sheet.cell(row=1, column=a+300).value = header_list_BE[a]


for data_BK in range(1, 10):
    for sheet_buku in range(1, 5):
        dt_tb = sheet_Buku_yang_pernah_ditulis.cell(row=1, column=sheet_buku).value
        header_list_BK.append(str(dt_tb) + "_" + str(abjad[data_BK-1]))
    for a in range(len(header_list_BK)):
        sheet.cell(row=1, column=a+345).value = header_list_BK[a]


for data_DK in range(1, 16):
    for sheet_diklat in range(1, 9):
        dt_tb = sheet_Diklat.cell(row=1, column=sheet_diklat).value
        header_list_DK.append(str(dt_tb) + "_" + str(abjad[data_DK-1]))
    for a in range(len(header_list_DK)):
        sheet.cell(row=1, column=a+381).value = header_list_DK[a]


for data_KT in range(1, 11):
    for sheet_karya_tulis in range(1, 6):
        dt_tb = sheet_Karya_Tulis.cell(row=1, column=sheet_karya_tulis).value
        header_list_KT.append(str(dt_tb) + "_" + str(abjad[data_KT-1]))
    for a in range(len(header_list_KT)):
        sheet.cell(row=1, column=a+501).value = header_list_KT[a]

# ============================================================================================================
for data_KS in range(1, 10):
    for sheet_kesejahteraan in range(1, 7):
        dt_tb = sheet_Kesejahteraan.cell(row=1, column=sheet_kesejahteraan).value
        header_list_KS.append(str(dt_tb) + "_" + str(abjad[data_KS-1]))
    for a in range(len(header_list_KS)):
        sheet.cell(row=1, column=a+551).value = header_list_KS[a]
# ===========================================================================================================

for data_TJ in range(1, 11):
    for sheet_tunjangan in range(1, 12):
        dt_tb = sheet_Tunjangan.cell(row=1, column=sheet_tunjangan).value
        header_list_TJ.append(str(dt_tb) + "_" + str(abjad[data_TJ-1]))
    for a in range(len(header_list_TJ)):
        sheet.cell(row=1, column=a+605).value = header_list_TJ[a]


for data_TT in range(1, 10):
    for sheet_tugas_tambahan in range(1, 5):
        dt_tb = sheet_Tugas_Tambahan.cell(row=1, column=sheet_tugas_tambahan).value
        header_list_TT.append(str(dt_tb) + "_" + str(abjad[data_TT-1]))
    for a in range(len(header_list_TT)):
        sheet.cell(row=1, column=a+715).value = header_list_TT[a]


for data_IN in range(1, 11):
    for sheet_inpassing in range(1, 8):
        dt_tb = sheet_Inpassing_Non_PNS.cell(row=1, column=sheet_inpassing).value
        header_list_IN.append(str(dt_tb) + "_" + str(abjad[data_IN-1]))
    for a in range(len(header_list_IN)):
        sheet.cell(row=1, column=a+751).value = header_list_IN[a]


for data_PH in range(1, 11):
    for sheet_penghargaan in range(1, 6):
        dt_tb = sheet_Penghargaan.cell(row=1, column=sheet_penghargaan).value
        header_list_PH.append(str(dt_tb) + "_" + str(abjad[data_PH-1]))
    for a in range(len(header_list_PH)):
        sheet.cell(row=1, column=a+821).value = header_list_PH[a]


for data_NT in range(1, 12):
    for sheet_nilai_tes in range(1, 7):
        dt_tb = sheet_Nilai_Tes.cell(row=1, column=sheet_nilai_tes).value
        header_list_NT.append(str(dt_tb) + "_" + str(abjad[data_NT-1]))
    for a in range(len(header_list_NT)):
        sheet.cell(row=1, column=a+871).value = header_list_NT[a]

for data_RG in range(1, 40):
    for sheet_riwayat_gaji_berkala in range(1, 8):
        dt_tb = sheet_Riwayat_Gaji_Berkala.cell(row=1, column=sheet_riwayat_gaji_berkala).value
        header_list_RG.append(str(dt_tb) + "_" + str(abjad[data_RG-1]))
    for a in range(len(header_list_RG)):
        sheet.cell(row=1, column=a+937).value = header_list_RG[a]


for data_RK in range(1, 13):
    for sheet_riwayat_karir in range(1, 13):
        dt_tb = sheet_Riwayat_Karir_Guru.cell(row=1, column=sheet_riwayat_karir).value
        header_list_RK.append(str(dt_tb) + "_" + str(abjad[data_RK-1]))
    for a in range(len(header_list_RK)):
        sheet.cell(row=1, column=a+1210).value = header_list_RK[a]
    

for data_RJPTK in range(1, 8):
    for sheet_riwayat_jabatan_p_tk in range(1, 4):
        dt_tb = sheet_Riwayat_Jabatan_P_TK.cell(row=1, column=sheet_riwayat_jabatan_p_tk).value
        header_list_RJPTK.append(str(dt_tb) + "_" + str(abjad[data_RJPTK-1]))
    for a in range(len(header_list_RJPTK)):
        sheet.cell(row=1, column=a+1354).value = header_list_RJPTK[a]

for data_KG in range(1, 23):
    for sheet_riwayat_kepangkatan_golongan in range(1, 7):
        dt_tb = sheet_Riwayat_Kepangkatan_Golongan.cell(row=1, column=sheet_riwayat_kepangkatan_golongan).value
        header_list_KG.append(str(dt_tb) + "_" + str(abjad[data_KG-1]))
    for a in range(len(header_list_KG)):
        sheet.cell(row=1, column=a+1375).value = header_list_KG[a]

for data_JF in range(1, 17):
    for sheet_riwayat_jabatan_fungsional in range(1, 4):
        dt_tb = sheet_Riwayat_Jabatan_Fungsional.cell(row=1, column=sheet_riwayat_jabatan_fungsional).value
        header_list_JF.append(str(dt_tb) + "_" + str(abjad[data_JF-1]))
    for a in range(len(header_list_JF)):
        sheet.cell(row=1, column=a+1507).value = header_list_JF[a]



print('Making Worksheet DONE...')



for x in range(len(ab)):
     data_list_RS = []
     data_list_RP = []
     data_list_KO = []
     data_list_AK = []
     data_list_BE = []
     data_list_BK = []
     data_list_DK = []
     data_list_KT = []
     data_list_KS = []
     data_list_TJ = []
     data_list_TT = []
     data_list_IN = []
     data_list_PH = []
     data_list_NT = []
     data_list_RG = []
     data_list_RK = []
     data_list_RJPTK = []
     data_list_KG = []
     data_list_JF = []
     sumber_datas = os.path.join(arr, 'Sumber', ab[x])
     wb = opx.load_workbook(sumber_datas, data_only=True)
     sheet_formulir_ptk = wb['Formulir_PTK']

     sheet_Riwayat_Sertifikasi=wb["Riwayat_Sertifikasi"]                          
     sheet_Riwayat_Pendidikan=wb["Riwayat_Pendidikan"]                           
     sheet_Kompetensi=wb["Kompetensi"]                            
     sheet_Anak=wb["Anak"]                            
     sheet_Beasiswa=wb["Beasiswa"]                            
     sheet_Buku_yang_pernah_ditulis=wb["Buku_yang_pernah_ditulis"]                            
     sheet_Diklat=wb["Diklat"]                            
     sheet_Karya_Tulis=wb["Karya_Tulis"]                          
     sheet_Kesejahteraan=wb["Kesejahteraan"]                          
     sheet_Tunjangan=wb["Tunjangan"]                          
     sheet_Tugas_Tambahan=wb["Tugas_Tambahan"]                            
     sheet_Inpassing_Non_PNS=wb["Inpassing_Non_PNS"]                          
     sheet_Penghargaan=wb["Penghargaan"]                          
     sheet_Nilai_Tes=wb["Nilai_Tes"]                          
     sheet_Riwayat_Gaji_Berkala=wb["Riwayat_Gaji_Berkala"]                            
     sheet_Riwayat_Karir_Guru=wb["Riwayat_Karir_Guru"]                            
     sheet_Riwayat_Jabatan_P_TK=wb["Riwayat_Jabatan_P_TK"]                            
     sheet_Riwayat_Kepangkatan_Golongan=wb["Riwayat_Kepangkatan_Golongan"]                            
     sheet_Riwayat_Jabatan_Fungsional=wb["Riwayat_Jabatan_Fungsional"]
     print(" "*100, end="\r")


     for formulir in range(len(list_header)):
         data = sheet_formulir_ptk.cell(row=formulir+1, column=6).value
         sheet.cell(row=x+2, column=formulir+1).value = data
        
    # RIWAYAT SERTIFIKASI
     for data_RS in range(1, 8):
         for sheet_riwayat_sertifikasi in range(1, 7):
             dt_tb = sheet_Riwayat_Sertifikasi.cell(row=data_RS+1, column=sheet_riwayat_sertifikasi).value
             data_list_RS.append(str(dt_tb))
         for a in range(len(data_list_RS)):
             sheet.cell(row=x+2, column=a+58).value = data_list_RS[a]

    # RIWAYAT PENDIDIKAN
     for data_RP in range(1, 11):
         for sheet_riwayat_pendidikan in range(1, 11):
             dt_tb = sheet_Riwayat_Pendidikan.cell(row=data_RP+1, column=sheet_riwayat_pendidikan).value
             data_list_RP.append(str(dt_tb))
         for a in range(len(data_list_RP)):
             sheet.cell(row=x+2, column=a+100).value = data_list_RP[a]

    # KOMPETENSI
     for data_KO in range(1, 9):
         for sheet_kompetensi in range(1, 3):
             dt_tb = sheet_Kompetensi.cell(row=data_KO+1, column=sheet_kompetensi).value
             data_list_KO.append(str(dt_tb))
         for a in range(len(data_list_KO)):
             sheet.cell(row=x+2, column=a+210).value = data_list_KO[a]

    # ANAK
     for data_AK in range(1, 9):
         for sheet_anak in range(1, 9):
             dt_tb = sheet_Anak.cell(row=data_AK+1, column=sheet_anak).value
             data_list_AK.append(str(dt_tb))
         for a in range(len(data_list_AK)):
             sheet.cell(row=x+2, column=a+228).value = data_list_AK[a]

    # BEASISWA
     for data_BE in range(1, 9):
         for sheet_beasiswa in range(1, 6):
             dt_tb = sheet_Beasiswa.cell(row=data_BE+1, column=sheet_beasiswa).value
             data_list_BE.append(str(dt_tb))
         for a in range(len(data_list_BE)):
             sheet.cell(row=x+2, column=a+300).value = data_list_BE[a]

    # BUKU YANG PERNAH DITULIS
     for data_BK in range(1, 9):
         for sheet_buku in range(1, 5):
             dt_tb = sheet_Buku_yang_pernah_ditulis.cell(row=data_BK+1, column=sheet_buku).value
             data_list_BK.append(str(dt_tb))
         for a in range(len(data_list_BK)):
             sheet.cell(row=x+2, column=a+345).value = data_list_BK[a]

    # DIKLAT
     for data_DK in range(1, 15):
         for sheet_diklat in range(1, 9):
             dt_tb = sheet_Diklat.cell(row=data_DK+1, column=sheet_diklat).value
             data_list_DK.append(str(dt_tb))
         for a in range(len(data_list_DK)):
             sheet.cell(row=x+2, column=a+381).value = data_list_DK[a]

    # KARYA TULIS
     for data_KT in range(1, 10):
         for sheet_karya_tulis in range(1, 6):
             dt_tb = sheet_Karya_Tulis.cell(row=data_KT+1, column=sheet_karya_tulis).value
             data_list_KT.append(str(dt_tb))
         for a in range(len(data_list_KT)):
             sheet.cell(row=x+2, column=a+501).value = data_list_KT[a]

    # KESEJAHTERAAN
     for data_KS in range(1, 9):
         for sheet_kesejahteraan in range(1, 7):
             dt_tb = sheet_Kesejahteraan.cell(row=data_KS+1, column=sheet_kesejahteraan).value
             data_list_KS.append(str(dt_tb))
         for a in range(len(data_list_KS)):
             sheet.cell(row=x+2, column=a+551).value = data_list_KS[a]

    # TUNJANGAN
     for data_TJ in range(1, 10):
         for sheet_tunjangan in range(1, 12):
             dt_tb = sheet_Tunjangan.cell(row=data_TJ+1, column=sheet_tunjangan).value
             data_list_TJ.append(str(dt_tb))
         for a in range(len(data_list_TJ)):
             sheet.cell(row=x+2, column=a+605).value = data_list_TJ[a]

    # TUGAS TAMBAHAN
     for data_TT in range(1, 9):
         for sheet_tugas_tambahan in range(1, 5):
             dt_tb = sheet_Tugas_Tambahan.cell(row=data_TT+1, column=sheet_tugas_tambahan).value
             data_list_TT.append(str(dt_tb))
         for a in range(len(data_list_TT)):
             sheet.cell(row=x+2, column=a+715).value = data_list_TT[a]

    # INPASSING NON PNS
     for data_IN in range(1, 10):
         for sheet_inpassing in range(1, 8):
             dt_tb = sheet_Inpassing_Non_PNS.cell(row=data_IN+1, column=sheet_inpassing).value
             data_list_IN.append(str(dt_tb))
         for a in range(len(data_list_IN)):
             sheet.cell(row=x+2, column=a+751).value = data_list_IN[a]

    # PENGHARGAAN
     for data_PH in range(1, 10):
         for sheet_penghargaan in range(1, 6):
             dt_tb = sheet_Penghargaan.cell(row=data_PH+1, column=sheet_penghargaan).value
             data_list_PH.append(str(dt_tb))
         for a in range(len(data_list_PH)):
             sheet.cell(row=x+2, column=a+821).value = data_list_PH[a]

    # NILAI TES
     for data_NT in range(1, 11):
         for sheet_nilai_tes in range(1, 7):
             dt_tb = sheet_Nilai_Tes.cell(row=data_NT+1, column=sheet_nilai_tes).value
             data_list_NT.append(str(dt_tb))
         for a in range(len(data_list_NT)):
             sheet.cell(row=x+2, column=a+871).value = data_list_NT[a]

    # RIWAYAT GAJI BERKALA
     for data_RG in range(1, 39):
         for sheet_riwayat_gaji_berkala in range(1, 8):
             dt_tb = sheet_Riwayat_Gaji_Berkala.cell(row=data_RG+1, column=sheet_riwayat_gaji_berkala).value
             data_list_RG.append(str(dt_tb))
         for a in range(len(data_list_RG)):
             sheet.cell(row=x+2, column=a+937).value = data_list_RG[a]

    # RIWAYAT KARIR
     for data_RK in range(1, 12):
         for sheet_riwayat_karir in range(1, 13):
             dt_tb = sheet_Riwayat_Karir_Guru.cell(row=data_RK+1, column=sheet_riwayat_karir).value
             data_list_RK.append(str(dt_tb))
         for a in range(len(data_list_RK)):
             sheet.cell(row=x+2, column=a+1210).value = data_list_RK[a]
        
    # RIWAYAT JABATAN PENDIDIK TENAGA KEPENDIDIKAN
     for data_RJPTK in range(1, 7):
         for sheet_riwayat_jabatan_p_tk in range(1, 4):
             dt_tb = sheet_Riwayat_Jabatan_P_TK.cell(row=data_RJPTK+1, column=sheet_riwayat_jabatan_p_tk).value
             data_list_RJPTK.append(str(dt_tb))
         for a in range(len(data_list_RJPTK)):
             sheet.cell(row=x+2, column=a+1354).value = data_list_RJPTK[a]

    # RIWAYAT KEPANGKATAN GOLONGAN
     for data_KG in range(1, 22):
         for sheet_riwayat_kepangkatan_golongan in range(1, 7):
             dt_tb = sheet_Riwayat_Kepangkatan_Golongan.cell(row=data_KG+1, column=sheet_riwayat_kepangkatan_golongan).value
             data_list_KG.append(str(dt_tb))
         for a in range(len(data_list_KG)):
             sheet.cell(row=x+2, column=a+1375).value = data_list_KG[a]

    # RIWAYAT JABATAN FUNGSIONAL
     for data_JF in range(1, 16):
         for sheet_riwayat_jabatan_fungsional in range(1, 4):
             dt_tb = sheet_Riwayat_Jabatan_Fungsional.cell(row=data_JF+1, column=sheet_riwayat_jabatan_fungsional).value
             data_list_JF.append(str(dt_tb))
         for a in range(len(data_list_JF)):
             sheet.cell(row=x+2, column=a+1507).value = data_list_JF[a]
            
     print("Diproses: ", x+1,".",ab[x])

penamaan = "hasil__"+tanggal_jam+".xlsx"
new_wr.save(penamaan)
new_wr.close()
print("DONE", penamaan)









